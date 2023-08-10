import openpyxl as op

class HomePageData:

    @staticmethod
    def getTestData(testcase_name):
        book = op.load_workbook("INSERT LOCAL MACHINE LOCATION FOR WORKBOOK WITH DATASET")
        sheet = book.active
        data_dict = {}
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == testcase_name:
                for column in range(2, sheet.max_column + 1):
                    data_dict[sheet.cell(row=1, column=column).value] = sheet.cell(row=row, column=column).value
        return [data_dict]
